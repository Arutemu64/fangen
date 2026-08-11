from typing import TYPE_CHECKING

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# Header styles
HEADER_ROW_FILL = PatternFill(
    start_color="70AD47", end_color="70AD47", fill_type="solid"
)
HEADER_ROW_FONT = Font(color="FFFFFF", bold=True)

# Topic styles
TOPIC_ROW_FONT = Font(color="000000", bold=True)

# Even style
EVEN_ROW_FILL = PatternFill(start_color="e2efda", end_color="e2efda", fill_type="solid")


def apply_final_formatting(
    ws: Worksheet,
    max_cell_length: int = 60,
    padding: int = 2,
    min_cell_length: int = 8,
    freeze_cell: str = "A2",  # Header
    *,
    auto_filter: bool = True,
) -> None:
    header_row = ws[1]
    for cell in header_row:
        cell.fill = HEADER_ROW_FILL
        cell.font = HEADER_ROW_FONT
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(col_idx)
        # Clamp the fitted width between a minimum (so short columns stay
        # comfortably readable and their header/filter arrow are not cramped)
        # and a maximum (so long free-text values wrap instead of stretching
        # the column off-screen).
        final_length = max_length + padding
        final_length = max(min_cell_length, min(final_length, max_cell_length))
        ws.column_dimensions[column_letter].width = final_length
    for row in ws.iter_rows():
        # Auto-height (updates height on Excel launch)
        ws.row_dimensions[row[0].row].height = None
        is_header = row[0].row == 1
        for cell in row:
            # Center the header for emphasis; left-align body cells because
            # long, wrapped Russian text is easier to read flush-left than
            # centered. Keep everything vertically centered within the row.
            cell.alignment = Alignment(
                wrapText=True,
                horizontal="center" if is_header else "left",
                vertical="center",
            )
    # Add filter dropdowns on the header row so users can sort and filter
    # columns directly. Skipped for sheets whose rows are not a flat table
    # (e.g. the plan, where topic rows act as section separators).
    if auto_filter:
        ws.auto_filter.ref = ws.dimensions
    # Freeze first row
    ws.freeze_panes = ws[freeze_cell]
