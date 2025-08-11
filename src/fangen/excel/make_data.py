import re
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from fangen.config import Config
from fangen.cosplay2.models.vo import RequestStatus
from fangen.db.repo import get_topics
from fangen.excel.formatting import apply_final_formatting, EVEN_ROW_FILL
from fangen.excel.utils import parse_request, check_excel_file
from rich import print


def make_data(filepath: Path, session: Session, config: Config) -> None:
    if filepath.exists():
        check_excel_file(filepath)

    wb = Workbook()
    wb.remove(wb.active)

    print("💻 Загружаем разделы и данные...")
    topics = get_topics(session)

    for topic in topics:
        print(f"""📄 Заполняем лист раздела '{topic.title}'...""")
        title = re.sub(r"[\/\\\?\*\:\[\]]", " ", topic.title)[:30]
        ws = wb.create_sheet(title=title)

        headers = [field.title for field in topic.fields]
        ws.append(headers)

        # Filter out approved
        requests = []
        for req in topic.requests:
            if req.status == RequestStatus.APPROVED:
                requests.append(req)

        for request in requests:
            current_row_index = requests.index(request) + 2
            request_context = parse_request(request, event_name=config.event_name)
            current_row = ws[current_row_index]
            for cell in current_row:
                header = headers[current_row.index(cell)]
                cell.value = str(request_context.get(header, ""))
                if current_row_index % 2 == 0:
                    cell.fill = EVEN_ROW_FILL

    print("🧹 Наводим красоту...")
    for ws in wb.worksheets:
        apply_final_formatting(ws)

    wb.save(filepath)
    print(f"💾 Файл сохранен по пути {filepath.absolute()}")
