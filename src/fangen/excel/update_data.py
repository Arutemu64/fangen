import csv
import re
from typing import TYPE_CHECKING

from openpyxl import Workbook
from rich import print

from fangen.common.data import parse_request
from fangen.cosplay2.models.vo import PlanNodeType
from fangen.db.repo import get_plan_nodes, get_topics_with_approved_requests
from fangen.excel.formatting import EVEN_ROW_FILL, apply_final_formatting
from fangen.excel.utils import check_excel_file

if TYPE_CHECKING:
    from collections.abc import Sequence
    from pathlib import Path

    from sqlalchemy.orm import Session

    from fangen.config import Config
    from fangen.db.models import Request

# Separator used when a single field holds several values. Rendering them as a
# joined string keeps both Excel and CSV readable instead of leaking Python's
# list repr (e.g. "['a', 'b']") into cells.
MULTI_VALUE_SEPARATOR = "; "

# Placeholder for empty cells in the Excel export. A visible dash reads better
# for humans than a blank cell. The CSV export deliberately uses a truly empty
# field instead (see make_data_csv), so tools like pandas parse it as NULL.
EXCEL_EMPTY = "—"


def render_value(value: object, empty: str = "") -> str:
    """Render a parsed request value as a single cell string.

    Multiple values are joined with MULTI_VALUE_SEPARATOR; missing or empty
    values become ``empty`` (a visible dash in Excel, a blank field in CSV).
    """
    if isinstance(value, list):
        parts = [str(item) for item in value if item is not None and str(item) != ""]
        return MULTI_VALUE_SEPARATOR.join(parts) if parts else empty
    if value is None or value == "":
        return empty
    return str(value)


def collect_summary(
    session: Session,
) -> tuple[list[str], Sequence[Request]]:
    """Return the flat summary layout: the union of all field titles across
    approved requests (prefixed with the ``info`` label column) and the
    requests themselves. Shared by the Excel summary sheet and the CSV export
    so both stay in sync.
    """
    plan_nodes = get_plan_nodes(session)
    requests = [
        node.request
        for node in plan_nodes
        if node.type is PlanNodeType.REQUEST and node.request is not None
    ]
    headers = ["info"]
    seen = {"info"}
    for request in requests:
        for value in request.values:
            if value.title not in seen:
                seen.add(value.title)
                headers.append(value.title)
    return headers, requests


def make_data(filepath: Path, session: Session, config: Config) -> None:
    if filepath.exists():
        check_excel_file(filepath)
    wb = Workbook()

    # Summary sheet with combined data
    print("💻 Заполняем сводный лист...")
    ws = wb.active
    ws.title = "Сводный"
    ws.freeze_panes = "B1"
    headers, requests = collect_summary(session)
    ws.append(headers)
    current_row_index = 2
    for request in requests:
        current_row = ws[current_row_index]
        data = parse_request(request)
        for cell, header in zip(current_row, headers, strict=False):
            if isinstance(header, str):
                cell.value = render_value(data.get(header), empty=EXCEL_EMPTY)
            if current_row_index % 2 == 0:
                cell.fill = EVEN_ROW_FILL
        current_row_index += 1
    print("🧹 Наводим красоту...")
    apply_final_formatting(ws, max_cell_length=config.max_cell_length, freeze_cell="B2")

    # One sheet per topic
    print("💻 Заполняем листы разделов...")
    topics = get_topics_with_approved_requests(session)
    for topic in topics:
        print(f"""📄 Заполняем лист раздела '{topic.title}'...""")
        title = re.sub(r"[\/\\\?\*\:\[\]]", " ", topic.title)[:30]
        ws = wb.create_sheet(title=title)

        headers = [field.title for field in topic.fields]
        ws.append(headers)

        for request in topic.requests:
            current_row_index = topic.requests.index(request) + 2
            data = parse_request(request)
            current_row = ws[current_row_index]
            for cell, header in zip(current_row, headers, strict=False):
                cell.value = render_value(data.get(header), empty=EXCEL_EMPTY)
                if current_row_index % 2 == 0:
                    cell.fill = EVEN_ROW_FILL
    print("🧹 Наводим красоту...")
    for ws in wb.worksheets[1:]:
        apply_final_formatting(ws, max_cell_length=config.max_cell_length)

    wb.save(filepath)
    print(f"💾 Файл сохранен по пути {filepath.absolute()}")


def make_data_csv(filepath: Path, session: Session) -> None:
    """Export the flat summary as a single CSV table for downstream AI/LLM use.

    Unlike the Excel workbook (which is styled for humans and split across
    per-topic sheets), this is one plain table: the union of every field, empty
    cells left blank, and multiple values joined into one string. That is the
    representation LLM tooling handles most reliably — a single table loads
    cleanly into pandas/DuckDB, and blank fields read as NULL rather than a
    literal dash. See update_data.py's module docstring notes for the rationale.
    """
    if filepath.exists():
        check_excel_file(filepath)
    print("💻 Экспортируем данные в CSV...")
    headers, requests = collect_summary(session)
    # utf-8-sig writes a BOM so Excel also opens the Cyrillic CSV correctly;
    # newline="" lets the csv module manage line endings and quoting.
    with filepath.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for request in requests:
            data = parse_request(request)
            writer.writerow([render_value(data.get(header)) for header in headers])
    print(f"💾 Файл сохранен по пути {filepath.absolute()}")
