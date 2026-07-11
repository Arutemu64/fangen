import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from rich import print
from rich.progress import track
from sqlalchemy.orm import Session

from fangen.common.data import get_node_data
from fangen.common.utils import format_template
from fangen.config import Config
from fangen.cosplay2.models.vo import PlanNodeType
from fangen.db.repo import get_plan_nodes
from fangen.excel.formatting import (
    EVEN_ROW_FILL,
    TOPIC_ROW_FONT,
    apply_final_formatting,
)
from fangen.excel.utils import check_excel_file

ALLOWED_PLAN_NODE_TYPES = [PlanNodeType.EVENT, PlanNodeType.TOPIC, PlanNodeType.REQUEST]


def make_plan(filepath: Path, session: Session, config: Config) -> None:
    if filepath.exists():
        check_excel_file(filepath)
        wb = load_workbook(filepath)
        print(f"💻 Открыт ранее созданный файл {filepath}")
    else:
        print(f"💻 Будет создан новый файл {filepath}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        basic_headers = ["{Инфо}"]
        ws.append(basic_headers)

    print("💻 Загружаем расписание и данные...")
    plan_nodes = get_plan_nodes(session)

    with config.dict_path.open(encoding="utf-8") as f:
        dictionary = json.load(f)

    for sheet in wb.worksheets:
        print(f"""📄 Обрабатываем лист '{sheet.title}'...""")
        first_row = sheet[1]
        headers = [cell.value for cell in first_row]
        print(f"🎩 Заголовки: {headers}")

        sheet.delete_rows(2, sheet.max_row - 1)

        current_row_index = 2
        request_number = 1
        nodes = [node for node in plan_nodes if node.type in ALLOWED_PLAN_NODE_TYPES]
        for node in track(nodes, "✏️ Заполняем строки..."):
            current_row = sheet[current_row_index]
            data = get_node_data(node)

            if node.type == PlanNodeType.REQUEST:
                data.update({"n": f"{request_number:03}"})
                request_number += 1

            for cell, header in zip(current_row, headers, strict=False):
                if isinstance(header, str):
                    cell.value = format_template(header, data, dictionary)
                if node.type is PlanNodeType.TOPIC:
                    cell.font = TOPIC_ROW_FONT
                if current_row_index % 2 == 0:
                    cell.fill = EVEN_ROW_FILL

            current_row_index += 1

    print("🧹 Наводим красоту...")
    for sheet in wb.worksheets:
        apply_final_formatting(sheet)

    wb.save(filepath)
    print(f"💾 Файл сохранен по пути {filepath.absolute()}")
