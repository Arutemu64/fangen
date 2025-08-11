from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fangen.cosplay2.models.plan import PlanNodeType
from fangen.db.models.node import PlanNode

MAX_CELL_LENGTH = 65
PADDING = 2

# Header styles
HEADER_ROW_FILL = PatternFill(
    start_color="70AD47", end_color="70AD47", fill_type="solid"
)
HEADER_ROW_FONT = Font(color="FFFFFF", bold=True)

# Topic styles
TOPIC_ROW_FONT = Font(color="000000", bold=True)

# Even style
EVEN_ROW_FILL = PatternFill(start_color="e2efda", end_color="e2efda", fill_type="solid")


def apply_row_style(node: PlanNode, row: tuple[Cell, ...], is_even: bool) -> None:
    for cell in row:
        if node.type is PlanNodeType.TOPIC:
            cell.font = TOPIC_ROW_FONT
        if is_even:
            cell.fill = EVEN_ROW_FILL


def apply_final_formatting(ws: Worksheet):
    header_row = ws[1]
    for cell in header_row:
        cell.fill = HEADER_ROW_FILL
        cell.font = HEADER_ROW_FONT
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(col_idx)
        final_length = max_length + PADDING
        if final_length < MAX_CELL_LENGTH:
            ws.column_dimensions[column_letter].width = final_length
        else:
            ws.column_dimensions[column_letter].width = MAX_CELL_LENGTH
    for row in ws.iter_rows():
        # Auto-height (updates height on Excel launch)
        ws.row_dimensions[row[0].row].height = None
        for cell in row:
            cell.alignment = Alignment(
                wrapText=True, horizontal="center", vertical="center"
            )
