from pathlib import Path

from openpyxl import load_workbook, Workbook
from rich import print
from rich.progress import track
from sqlalchemy.orm import Session

from fangen.config import Config
from fangen.cosplay2.models.plan import PlanNodeType
from fangen.db.repo import get_plan_nodes
from fangen.plan.utils import get_node_context, format_header
from fangen.plan.formatting import (
    apply_final_formatting,
    apply_row_style,
)

ALLOWED_PLAN_NODE_TYPES = [PlanNodeType.EVENT, PlanNodeType.TOPIC, PlanNodeType.REQUEST]


def check_excel_file(filepath: Path):
    try:
        with open(filepath, "a"):
            pass
    except PermissionError:
        raise PermissionError("Закройте Excel-файл для обновления")


def make_plan(filepath: Path, session: Session, config: Config) -> None:
    if filepath.exists():
        check_excel_file(filepath)
        wb = load_workbook(filepath)
        print(f"💻 Открыт ранее созданный файл {filepath}")
    else:
        print(f"💻 Будет создан новый файл {filepath}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
        basic_headers = ["{Инфо}"]
        ws.append(basic_headers)

    plan_nodes = get_plan_nodes(session)

    for sheet in wb.worksheets:
        print(f"""📄 Обрабатываем лист '{sheet.title}'...""")
        first_row = sheet[1]
        headers = [cell.value for cell in first_row]
        print(f"🎩 Заголовки: {headers}")

        sheet.delete_rows(2, sheet.max_row - 1)

        current_row_index = 2
        request_number = 1
        for node in track(plan_nodes, "✏️ Заполняем строки..."):
            if node.type in ALLOWED_PLAN_NODE_TYPES:
                current_row = sheet[current_row_index]
                context = get_node_context(
                    node, request_number, event_name=config.event_name
                )

                for cell in current_row:
                    header = headers[current_row.index(cell)]
                    if isinstance(header, str):
                        cell.value = format_header(
                            header, context, dict_path=config.dict_path
                        )

                if node.type == PlanNodeType.REQUEST:
                    request_number += 1

                apply_row_style(node, current_row, (current_row_index % 2 == 0))
                current_row_index += 1

        print("🧹 Наводим красоту...")
        apply_final_formatting(sheet)

    wb.save(filepath)
    print(f"💾 Файл сохранен по пути {filepath.absolute()}")
